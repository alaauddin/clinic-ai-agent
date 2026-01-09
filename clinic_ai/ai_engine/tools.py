from langchain.tools import tool
from clinic_ai.models import Doctor, ClinicInfo, Appointment, Clinic, DoctorAvailability
from django.db.models import Q
from datetime import datetime
from clinic_ai.context import current_user
from django.conf import settings
import os
import uuid
import openpyxl
from openpyxl.styles import Font as XLFont
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image, PageTemplate, BaseDocTemplate, Frame
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.lib.units import cm, inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import arabic_reshaper
from bidi.algorithm import get_display

# Register Arabic Fonts
try:
    pdfmetrics.registerFont(TTFont('Arabic', '/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf'))
    pdfmetrics.registerFont(TTFont('Arabic-Bold', '/usr/share/fonts/truetype/noto/NotoNaskhArabic-Bold.ttf'))
except:
    pass

def fix_arabic(text):
    if not text:
        return ""
    reshaped_text = arabic_reshaper.reshape(text)
    bidi_text = get_display(reshaped_text)
    return bidi_text

@tool
def list_clinics(query: str = ""):
    """
    List all available clinics and their doctors in the medical center.
    STRICT RULE: Only return clinics found in the database. Never hallucinate or add fictional clinics.
    """
    clinics = Clinic.objects.all().prefetch_related('doctors')
    if not clinics.exists():
        return "لا توجد عيادات مسجلة حالياً."
    
    table = "| # | العيادة | الموقع | الأطباء |\n"
    table += "| :--- | :--- | :--- | :--- |\n"
    
    for i, c in enumerate(clinics):
        doctors = c.doctors.all()
        doc_list = "<br>".join([f"• {d.name} ({d.specialty})" for d in doctors]) if doctors.exists() else "لا يوجد أطباء حالياً"
        table += f"| {i+1} | {c.name} | {c.location} | {doc_list} |\n"
    
    return table

@tool
def list_all_doctors(query: str = ""):
    """
    استرجاع قائمة بجميع الأطباء في المركز الطبي مع تخصصاتهم وعياداتهم.
    استخدم هذه الأداة عندما يطلب المستخدم تقريراً أو قائمة عامة لجميع الأطباء.
    """
    docs = Doctor.objects.all().select_related('clinic')
    if not docs.exists():
        return "لا يوجد أطباء مسجلون حالياً."
    
    results = []
    for d in docs:
        results.append({
            "اسم الطبيب": d.name,
            "التخصص": d.specialty,
            "العيادة": d.clinic.name if d.clinic else "غير محدد"
        })
    import json
    return json.dumps(results, ensure_ascii=False)

@tool
def get_doctor_availability(doctor_info: str):
    """
    Search for doctor availability. 
    You can search by: name (e.g., 'Dr. Ahmed'), specialty (e.g., 'Dermatology'), or clinic name.
    """
    parts = [p.strip() for p in doctor_info.replace('،', ',').split(',')]
    query = parts[0]
    clinic_name = parts[1] if len(parts) > 1 else None

    # Search Logic
    words = query.split()
    q_obj = Q()
    for word in words:
        clean_word = word[2:] if word.startswith('ال') and len(word) > 3 else word
        q_obj &= (Q(name__icontains=clean_word) | Q(specialty__icontains=clean_word) | Q(clinic__name__icontains=clean_word))
    
    doctors = Doctor.objects.filter(q_obj).select_related('clinic')
    if clinic_name:
        doctors = doctors.filter(clinic__name__icontains=clinic_name)

    if not doctors.exists():
        return "لا يوجد أطباء بهذا الوصف حالياً."
    
    from datetime import datetime, timedelta
    now = datetime.now()
    table_rows = []

    for doc in doctors:
        availabilities = doc.availabilities.all()
        for i in range(7):
            check_date = (now + timedelta(days=i)).date()
            day_val = check_date.weekday()
            
            day_slots = []
            for avail in availabilities.filter(day_of_week=day_val):
                curr_dt = datetime.combine(check_date, avail.start_time)
                end_dt = datetime.combine(check_date, avail.end_time)
                
                while curr_dt < end_dt:
                    slot_start = curr_dt
                    # Exact 30-min window check
                    is_booked = Appointment.objects.filter(
                        doctor=doc,
                        appointment_date__range=(slot_start - timedelta(seconds=1799), slot_start + timedelta(seconds=1799)),
                        status__in=['pending', 'confirmed']
                    ).exists()
                    
                    if not is_booked and curr_dt > now:
                        day_slots.append(curr_dt.strftime('%H:%M'))
                    curr_dt += timedelta(minutes=30)
            
            if day_slots:
                days_ar = ["الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]
                day_name = days_ar[day_val]
                clinic_label = f" ({doc.clinic.name})" if doc.clinic else ""
                table_rows.append(f"| {day_name} | {check_date} | {doc.name}{clinic_label} | {', '.join(day_slots)} |")

    if not table_rows:
        return "عذراً، لا توجد مواعيد متاحة لهؤلاء الأطباء في الأيام القادمة."

    table = "| اليوم | التاريخ | الطبيب (العيادة) | الأوقات المتاحة |\n"
    table += "| :--- | :--- | :--- | :--- |\n"
    table += "\n".join(table_rows)
    
    return table

@tool
def get_clinic_general_info(query: str = ""):
    """Get general clinic information like working hours, location, and phone."""
    info = ClinicInfo.objects.first()
    if not info:
        return "لا تتوفر معلومات عامة عن العيادة حالياً."
    
    return f"ساعات العمل: {info.working_hours}\nالموقع: {info.location}\nالهاتف: {info.phone}"

@tool
def get_available_doctors_by_date(date_str: str):
    """
    استرجاع قائمة بالأطباء المتاحين في تاريخ معين، مع عرض الأوقات المتاحة لكل طبيب.
    المدخل: التاريخ بصيغة 'YYYY-MM-DD'.
    """
    try:
        from datetime import datetime, time, timedelta
        from django.utils import timezone
        search_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        day_val = search_date.weekday()
        
        # Get all availabilities for this day of week
        availabilities = DoctorAvailability.objects.filter(day_of_week=day_val).select_related('doctor')
        
        if not availabilities.exists():
            return f"لا يوجد أطباء متاحون في هذا التاريخ ({date_str})."
        
        # Get existing appointments for this date to check for overlaps
        booked_appointments = Appointment.objects.filter(
            appointment_date__date=search_date,
            status__in=['pending', 'confirmed']
        ).values_list('doctor_id', 'appointment_date')
        
        # Group booked times by doctor (handle awareness)
        booked_map = {}
        for doc_id, appt_time in booked_appointments:
            if doc_id not in booked_map:
                booked_map[doc_id] = []
            # Ensure we compare times in the same context (UTC since settings.TIME_ZONE='UTC')
            booked_map[doc_id].append(appt_time.time())

        results = []
        for avail in availabilities:
            doctor = avail.doctor
            slots = []
            curr_time = datetime.combine(search_date, avail.start_time)
            end_datetime = datetime.combine(search_date, avail.end_time)
            
            while curr_time < end_datetime:
                slot_time = curr_time.time()
                # Overlap logic: A slot at 'slot_time' is blocked if any appointment exists that starts
                # within 29 minutes and 59 seconds before or after this time (the '30-minute rule').
                # This ensures back-to-back appointments (e.g., 10:00 and 10:30) are allowed, but anything else is blocked.
                is_booked = any(
                    abs((datetime.combine(search_date, slot_time) - appt_dt.replace(tzinfo=None)).total_seconds()) < 1800 
                    for d_id, appt_dt in booked_appointments if d_id == doctor.id
                )
                
                if not is_booked:
                    # Return 24h format for better AI parsing, maybe with emoji
                    slots.append(slot_time.strftime('%H:%M'))
                
                curr_time += timedelta(minutes=30)

            if slots:
                results.append(f"{len(results)+1}. الطبيب: {doctor.name} ({doctor.specialty})\n   الأوقات المتاحة: {', '.join(slots)}")
        
        if not results:
            return f"جميع المواعيد محجوزة في هذا التاريخ ({date_str})."
        
        return "\n\n".join(results)
    except Exception as e:
        return f"حدث خطأ أثناء البحث عن المواعيد المتاحة: {str(e)}"

@tool
def get_upcoming_availability_for_clinic(clinic_name: str):
    """
    استرجاع كافة المواعيد المتاحة (الغير محجوزة) لجميع الأطباء في عيادة معينة للأيام السبعة القادمة.
    استخدم هذه الأداة فور اختيار المستخدم للعيادة لتعرض له الخيارات المتاحة مباشرة.
    """
    try:
        from datetime import datetime, time, timedelta
        from django.utils import timezone
        
        clinic = Clinic.objects.filter(name__icontains=clinic_name).first()
        if not clinic:
            return f"العيادة '{clinic_name}' غير موجودة."
        
        doctors = clinic.doctors.all()
        if not doctors.exists():
            return f"لا يوجد أطباء مسجلون في عيادة {clinic.name} حالياً."
        
        now = datetime.now()
        table_rows = []
        
        for doctor in doctors:
            availabilities = doctor.availabilities.all()
            for i in range(7):
                check_date = (now + timedelta(days=i)).date()
                day_val = check_date.weekday()
                
                day_slots = []
                for avail in availabilities.filter(day_of_week=day_val):
                    curr_time = datetime.combine(check_date, avail.start_time)
                    end_datetime = datetime.combine(check_date, avail.end_time)
                    
                    while curr_time < end_datetime:
                        slot_time = curr_time.time()
                        # Range check for 30-minute overlap.
                        # We use 29.9 minutes to allow exact back-to-back bookings while catching any internal overlaps.
                        from datetime import timedelta
                        slot_start = datetime.combine(check_date, slot_time)
                        is_booked = Appointment.objects.filter(
                            doctor=doctor,
                            appointment_date__range=(slot_start - timedelta(seconds=1799), slot_start + timedelta(seconds=1799)),
                            status__in=['pending', 'confirmed']
                        ).exists()
                        
                        is_future = datetime.combine(check_date, slot_time) > now
                        if not is_booked and is_future:
                            day_slots.append(slot_time.strftime('%H:%M'))
                        curr_time += timedelta(minutes=30)
                
                if day_slots:
                    days_ar = ["الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]
                    day_name = days_ar[day_val]
                    table_rows.append(f"| {day_name} | {check_date} | {doctor.name} | {', '.join(day_slots)} |")
        
        if not table_rows:
            return f"عذراً، لا توجد مواعيد متاحة في عيادة {clinic.name} خلال الـ 7 أيام القادمة."
        
        table = "| اليوم | التاريخ | الطبيب | الأوقات المتاحة |\n"
        table += "| :--- | :--- | :--- | :--- |\n"
        table += "\n".join(table_rows)
        
        return table
    except Exception as e:
        return f"خطأ في جلب المواعيد: {str(e)}"

@tool
def book_appointment(appointment_info: str):
    """
    حجز موعد جديد للمريض. 
    **🚨 متطلب إلزامي جداً 🚨**: يجب طلب جميع بيانات المريض (الاسم، تاريخ الميلاد YYYY-MM-DD، الهاتف، البريد) **في كل جلسة دردشة جديدة**. 
    لا تعتمد أبداً على بيانات من سجلات سابقة؛ اطلبها من المستخدم مباشرة في كل مرة يطلب فيها حجزاً جديداً.
    المدخل يجب أن يكون: 'اسم العيادة، اسم الطبيب، التاريخ والوقت YYYY-MM-DD HH:MM، اسم المريض، تاريخ الميلاد، الهاتف، البريد'.
    """
    user = current_user.get()
    if user is None or not user.is_authenticated:
        return "يجب عليك تسجيل الدخول أولاً لحجز موعد."
    
    try:
        from django.utils import timezone
        from datetime import datetime
        parts = [p.strip() for p in appointment_info.replace('،', ',').split(',')]
        if len(parts) < 7:
            return "بيانات ناقصة. المطلوب: (العيادة، الطبيب، الموعد YYYY-MM-DD HH:MM، الاسم، تاريخ الميلاد، الهاتف، البريد)."
        
        cl_name, doc_name, date_str, p_name, p_dob_str, p_phone, p_email = parts[:7]

        clinic = Clinic.objects.filter(name__icontains=cl_name).first()
        if not clinic: return f"العيادة '{cl_name}' غير موجودة."

        doctor = Doctor.objects.filter(clinic=clinic, name__icontains=doc_name).first()
        if not doctor: return f"الطبيب '{doc_name}' غير موجود في هذه العيادة."
        
        try:
            appt_date_naive = datetime.strptime(date_str, '%Y-%m-%d %H:%M')
            # Make it aware based on settings (UTC)
            appt_date = timezone.make_aware(appt_date_naive, timezone.get_current_timezone())
        except Exception:
            return "تنسيق التاريخ والوقت غير صحيح. استخدم YYYY-MM-DD HH:MM."

        if appt_date <= timezone.now():
            return "الموعد يجب أن يكون في المستقبل."

        # Check availability
        day_val = appt_date.weekday()
        time_val = appt_date.time()
        if not DoctorAvailability.objects.filter(doctor=doctor, day_of_week=day_val, start_time__lte=time_val, end_time__gt=time_val).exists():
            return "الطبيب غير متاح في هذا الوقت بناءً على جدوله الأسبوعي."

        # Collision Protection (30-Minute Window)
        from datetime import timedelta
        # Ensure no other appointment overlaps with this 30-minute block
        # Buffer of 29 minutes and 59 seconds ensures back-to-back slots are allowed
        start_search = appt_date - timedelta(seconds=1799)
        end_search = appt_date + timedelta(seconds=1799)
        if Appointment.objects.filter(doctor=doctor, appointment_date__range=(start_search, end_search), status__in=['pending', 'confirmed']).exists():
            return f"عذراً، هذا الموعد {date_str} متداخل مع موعد آخر (مدة الموعد 30 دقيقة). يرجى اختيار وقت آخر."

        # Create
        try:
            p_dob = datetime.strptime(p_dob_str, '%Y-%m-%d').date()
        except:
            return "تنسيق تاريخ الميلاد غير صحيح (YYYY-MM-DD)."

        appointment = Appointment.objects.create(
            user=user, clinic=clinic, doctor=doctor, appointment_date=appt_date,
            patient_name=p_name, patient_dob=p_dob, patient_phone=p_phone, patient_email=p_email
        )
        
        # Notifications (Send silently or check settings)
        try:
            from django.core.mail import send_mail
            from django.conf import settings
            subject = f"تأكيد حجز - {clinic.name}"
            msg = f"تم الحجز بنجاح!\nالطبيب: {doctor.name}\nالموعد: {date_str}\nالاسم: {p_name}\nرقم الحجز: {appointment.id}"
            send_mail(subject, msg, settings.DEFAULT_FROM_EMAIL, [p_email])
        except:
            pass

        return f"تهانينا! تم حجز موعدك بنجاح. رقم الحجز الخاص بك هو {appointment.id}. تم إرسال رسالة تأكيد إلكترونية إلى {p_email}."
    except Exception as e:
        return f"خطأ تقني: {str(e)}"

@tool
def list_user_appointments(query: str = ""):
    """List appointments for the current user."""
    user = current_user.get()
    if user is None or not user.is_authenticated:
        return "يجب عليك تسجيل الدخول أولاً لعرض مواعيدك."
    
    appts = Appointment.objects.filter(user=user).order_by('-appointment_date')
    if not appts.exists():
        return "ليس لديك أي مواعيد محجوزة حالياً."
    
    results = []
    for appt in appts:
        results.append(f"موعد #{appt.id}: {appt.appointment_date} - {appt.doctor.name} - الحالة: {appt.get_status_display()}")
    
    return "\n".join(results)

@tool
def generate_excel_report(data_json: str):
    """
    إنشاء ملف Excel من البيانات المقدمة.
    يجب أن تكون المدخلات عبارة عن JSON يمثل قائمة من القواميس (List of Dictionaries).
    مثال: '[{"اسم المريض": "أحمد", "الموعد": "2026-01-01"}]'
    ستقوم هذه الأداة بحفظ الملف وإرجاع رابط التحميل.
    """
    import json
    try:
        data = json.loads(data_json)
        if not data or not isinstance(data, list):
            return "يجب أن تكون البيانات قائمة من القواميس."
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Extract headers from the first row
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = XLFont(bold=True)
            
        # Add data rows
        for row_num, entry in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=str(entry.get(header, "")))
        
        # Save file to media root
        filename = f"report_{uuid.uuid4().hex[:8]}.xlsx"
        filepath = os.path.join(settings.MEDIA_ROOT, filename)
        wb.save(filepath)
        
        file_url = f"{settings.MEDIA_URL}{filename}"
        return f"تم إنشاء ملف Excel بنجاح. يمكنك تحميله من الرابط التالي: {file_url}"
    except Exception as e:
        return f"حدث خطأ أثناء إنشاء ملف Excel: {str(e)}"

@tool
def generate_pdf_report(data_json: str):
    """
    إنشاء ملف PDF استثنائي واحترافي بتصميم Dashboard حديث.
    يجب أن تكون المدخلات عبارة عن JSON يمثل قائمة من القواميس.
    """
    import json
    try:
        data = json.loads(data_json)
        if not data or not isinstance(data, list):
            return "يجب أن تكون البيانات قائمة من القواميس."
        
        filename = f"premium_report_{uuid.uuid4().hex[:8]}.pdf"
        filepath = os.path.join(settings.MEDIA_ROOT, filename)
        logo_path = os.path.join(settings.MEDIA_ROOT, 'assets/logo.png')
        
        # Color Palette - Premium Navy & Cyan
        NAVY = colors.HexColor("#0F172A")
        LIGHT_NAVY = colors.HexColor("#1E293B")
        CYAN = colors.HexColor("#38BDF8")
        BG_LIGHT = colors.HexColor("#F8FAFC")
        
        class PremiumDoc(BaseDocTemplate):
            def __init__(self, filename, **kw):
                super().__init__(filename, **kw)
                frame = Frame(self.leftMargin, self.bottomMargin, self.width, self.height - 100, id='normal')
                self.addPageTemplates([PageTemplate(id='First', frames=frame, onPage=self.on_page)])

            def on_page(self, canvas, doc):
                canvas.saveState()
                # Background color for header
                canvas.setFillColor(NAVY)
                canvas.rect(0, A4[1]-120, A4[0], 120, fill=1)
                
                # Draw Logo if exists
                if os.path.exists(logo_path):
                    canvas.drawImage(logo_path, 40, A4[1]-100, width=80, height=80, mask='auto')
                
                # Header Text
                canvas.setFillColor(colors.white)
                canvas.setFont('Arabic-Bold', 22)
                canvas.drawRightString(A4[0]-40, A4[1]-60, fix_arabic("المركز الطبي الذكي"))
                canvas.setFont('Arabic', 10)
                canvas.drawRightString(A4[0]-40, A4[1]-85, fix_arabic("Smart Clinic Center - Premium AI Intelligence"))
                
                # Bottom Decorative Line
                canvas.setStrokeColor(CYAN)
                canvas.setLineWidth(3)
                canvas.line(40, A4[1]-120, A4[0]-40, A4[1]-120)
                
                # Footer
                canvas.setFillColor(colors.grey)
                canvas.setFont('Arabic', 9)
                canvas.drawString(40, 20, fix_arabic(f"تاريخ الإصدار: {datetime.now().strftime('%Y-%m-%d')}"))
                canvas.drawRightString(A4[0]-40, 20, fix_arabic(f"صفحة {canvas.getPageNumber()}"))
                canvas.restoreState()

        doc = PremiumDoc(filepath, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=140, bottomMargin=40)
        elements = []
        styles = getSampleStyleSheet()
        
        # Dashboard Style Title
        title_style = ParagraphStyle(
            'TitleDash',
            fontName='Arabic-Bold',
            fontSize=16,
            textColor=LIGHT_NAVY,
            spaceAfter=30,
            alignment=TA_RIGHT
        )
        elements.append(Paragraph(fix_arabic("تقرير تحليل البيانات والفرق الطبية"), title_style))
        
        # Summary Area (Mini Cards)
        summary_data = [
            [
                Paragraph(fix_arabic(f"إجمالي السجلات: {len(data)}"), ParagraphStyle('S1', fontName='Arabic', fontSize=12, textColor=NAVY)),
                Paragraph(fix_arabic("الحالة: تقرير رسمي"), ParagraphStyle('S2', fontName='Arabic', fontSize=12, textColor=colors.HexColor("#10B981"))) # Emerald
            ]
        ]
        summary_table = Table(summary_data, colWidths=[150, 150])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F1F5F9")),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#E2E8F0")),
            ('PADDING', (0,0), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 40))
        
        # Main Data Table
        headers = [fix_arabic(h) for h in list(data[0].keys())]
        table_data = [headers]
        for entry in data:
            row = [fix_arabic(str(entry.get(h, ""))) for h in list(data[0].keys())]
            table_data.append(row)
            
        main_table = Table(table_data, hAlign='CENTER', repeatRows=1)
        main_table.setStyle(TableStyle([
            # Modern Header
            ('BACKGROUND', (0, 0), (-1, 0), LIGHT_NAVY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Arabic-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Subtle Body
            ('FONTNAME', (0, 1), (-1, -1), 'Arabic'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ('GRID', (0, 0), (-1, -1), 0.1, colors.HexColor("#CBD5E1")),
            ('LINEBELOW', (0, 0), (-1, 0), 2, CYAN),
        ]))
        
        elements.append(main_table)
        elements.append(Spacer(1, 50))
        
        # Stamp / Signature Area
        stamp_style = ParagraphStyle('Stamp', fontName='Arabic', fontSize=10, textColor=colors.lightgrey, alignment=TA_CENTER)
        elements.append(Paragraph(fix_arabic("تمت المصادقة الرقمية بواسطة نظام ذكاء المركز الطبي"), stamp_style))
        
        doc.build(elements)
        
        file_url = f"{settings.MEDIA_URL}{filename}"
        return f"تم إنشاء تقرير PDF استثنائي بنجاح. يمكنك تحميله من الرابط التالي: {file_url}"
    except Exception as e:
        import traceback
        return f"حدث خطأ أثناء إنشاء ملف PDF الاستثنائي: {str(e)}\n{traceback.format_exc()}"
